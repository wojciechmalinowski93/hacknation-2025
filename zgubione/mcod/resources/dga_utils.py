import csv
import datetime
import logging
import os
import uuid
from io import BytesIO
from mimetypes import guess_extension, guess_type
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

import numpy as np
import pandas as pd
import requests
import sentry_sdk
from cache_memoize import cache_memoize
from celery import states
from chardet import detect as detect_encoding
from django.apps import apps
from django.conf import settings
from django.core.cache import caches
from django.core.exceptions import MultipleObjectsReturned, ObjectDoesNotExist
from django.core.files.uploadedfile import InMemoryUploadedFile, SimpleUploadedFile
from django.db import transaction
from django.db.models import QuerySet
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mcod.core.utils import clean_columns_in_dataframe, save_df_to_xlsx
from mcod.resources.dga_constants import (
    ALLOWED_DGA_INSTITUTIONS,
    DGA_COLUMNS,
    DGA_RESOURCE_EXTENSIONS,
)
from mcod.resources.exceptions import FailedValidationException, PendingValidationException
from mcod.resources.goodtables_checks import ZERO_DATA_ROWS_MSG

logger = logging.getLogger("mcod")


def get_main_dga_resource() -> Optional["Resource"]:  # noqa: F821
    from mcod.resources.models import AggregatedDGAInfo

    dga_info: Optional[AggregatedDGAInfo] = AggregatedDGAInfo.objects.last()
    return dga_info.main_dga_resource if dga_info else None


def get_main_dga_dataset() -> Optional["Dataset"]:  # noqa: F821
    from mcod.resources.models import AggregatedDGAInfo

    dga_info: Optional[AggregatedDGAInfo] = AggregatedDGAInfo.objects.last()
    return dga_info.main_dga_dataset if dga_info else None


def create_df_from_dga_file(
    file: Union[InMemoryUploadedFile, BytesIO],
    extension: Optional[str],
) -> Optional[pd.DataFrame]:
    """
    Function returns DataFrame for given file and extension or returns
    None if extension is not one of: `csv`, `xls`, `xlsx`.
    """
    if extension in ("xls", "xlsx"):
        return pd.read_excel(file)

    if extension == "csv":
        raw_data = file.read()
        file.seek(0)
        result = detect_encoding(raw_data)
        encoding = result["encoding"]
        if encoding is None:
            logger.error("Could not detect dga file encoding")
            return None

        content = raw_data.decode(encoding)
        dialect = csv.Sniffer().sniff(content)
        return pd.read_csv(file, dialect=dialect)


def validate_dga_df_columns(df: pd.DataFrame) -> bool:
    """Returns True if DataFrame columns are correctly named and ordered, False otherwise."""
    return df.columns.to_list() == DGA_COLUMNS


def validate_dga_file_columns(
    file: Union[InMemoryUploadedFile, BytesIO],
    extension: Optional[str],
) -> bool:
    """
    Returns True if file columns are correctly named and ordered (according to DGA file rules),
    False otherwise.
    """
    # Create DataFrame for given file and extension
    try:
        df: Optional[pd.DataFrame] = create_df_from_dga_file(file, extension)
    except Exception as e:
        logger.exception(f"Error reading dga file: {e}")
        return False

    # Return False if DataFrame could not be created
    if df is None:
        return False

    # Return True if DataFrame columns are ok, False otherwise
    return validate_dga_df_columns(df)


def get_dga_resource_for_institution(
    organization_id: Union[int, str],
    exclude_resource_id: Optional[Union[int, str]] = None,
) -> Optional["Resource"]:  # noqa: F821
    """
    Returns DGA Resource object for given Organization. Main DGA Resource
    and Resource object with given PK are excluded.
    """
    from mcod.resources.models import Resource

    # Exclude main DGA Resource and Resource with given id
    exclude_objects_ids: List[Union[int, str]] = []
    main_dga_resource: Optional[Resource] = get_main_dga_resource()

    if main_dga_resource:
        exclude_objects_ids.append(main_dga_resource.pk)

    if exclude_resource_id:
        exclude_objects_ids.append(exclude_resource_id)

    query = Resource.objects.filter(
        dataset__organization=organization_id,
        contains_protected_data=True,
        status="published",
    ).exclude(id__in=exclude_objects_ids)

    count_dga_resources: int = query.count()
    if count_dga_resources > 1:
        error_message = f"Found {count_dga_resources} DGA Resources for organization: " f"{organization_id}"
        logger.error(error_message)
        raise MultipleObjectsReturned(error_message)

    return query.first()


def create_uploaded_file_from_path(path: str) -> Union[SimpleUploadedFile, None]:
    try:
        with open(path, "rb") as tmp_file:
            file_content = tmp_file.read()

        file_name = os.path.basename(path)
        content_type, _ = guess_type(path)
        file = SimpleUploadedFile(
            name=file_name,
            content=file_content,
            content_type=content_type,
        )
    except FileNotFoundError:
        logger.error(f"Cannot find a file at {path}")
        return None
    except Exception as e:
        logger.error(f"Exception while uploading from path {path} occurred: {e}")
        return None
    return file


def save_temp_dga_file(file: InMemoryUploadedFile) -> str:
    file_name = file.name
    base_name, extension = os.path.splitext(file_name)
    unique_id = uuid.uuid4()
    temp_file_name = f"{base_name}_{unique_id}{extension}"
    directory = settings.DGA_RESOURCE_CREATION_STAGING_ROOT
    if not os.path.exists(directory):
        os.makedirs(directory)
    file_path = os.path.join(directory, temp_file_name)

    with open(file_path, "wb+") as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    return temp_file_name


def key_generator_for_create_main_xlsx_file(*args, **kwargs) -> str:
    return "main_xlsx_file_path_to_clean"


def get_or_create_main_dga_path() -> Path:
    directory: Path = Path(settings.MAIN_DGA_RESOURCE_XLSX_CREATION_ROOT)
    if not os.path.exists(directory):
        os.makedirs(directory)

    today_date: str = datetime.datetime.now().strftime("%Y%m%d")
    file_name: str = f"{settings.MAIN_DGA_XLSX_FILE_NAME_PREFIX} {today_date}.xlsx"
    return directory / file_name


def get_all_dga_resources_sorted_by_organizations() -> QuerySet:
    Resource = apps.get_model("resources", "Resource")

    # Exclude current main DGA Resource in file creation process.
    main_dga_resource: Optional[Resource] = get_main_dga_resource()
    main_dga_id = main_dga_resource.pk if main_dga_resource else None

    dga_resources: QuerySet = (
        Resource.objects.filter(contains_protected_data=True, status="published")
        .exclude(id=main_dga_id)
        .select_related("dataset__organization")
        .order_by("dataset__organization__title")
    )
    return dga_resources


def get_ckan_dga_resource_df(resource: "Resource") -> Optional[pd.DataFrame]:  # noqa: F821
    """
    Creates df (DataFrame) for DGA Resource that is harvested by CKAN.
    Returns the df based on currently available remote data if DGA compatible
    or None otherwise.
    """
    # Link to remote data where CKAN harvested data should be located
    url: Optional[str] = resource.link
    if url is None:
        logger.error(f"CKAN Resource has no link. Resource id: {resource.pk}")
        return None

    # Try to fetch the data
    try:
        response: requests.models.Response = request_remote_dga(url)
    except requests.exceptions.RequestException:
        logger.exception(f"Site not responding. Resource id: {resource.pk}; url: {url}")
        return None

    # Check response status code
    status_code: int = response.status_code
    if not status_code == 200:
        logger.error(f"Status code not 200: {status_code}. Resource id: {resource.pk}; url: {url}")
        return None

    # Check if file extension is correct
    extension_for_remote: Optional[str] = get_remote_extension_if_correct_dga_content_type(response)
    if extension_for_remote is None:
        logger.error(f"Incorrect file extension. Resource id: {resource.pk}; url: {url}")
        return None

    # Convert data to DataFrame
    file_data = BytesIO(response.content)
    df: Optional[pd.DataFrame] = create_df_from_dga_file(file_data, extension_for_remote)
    if df is None:
        logger.error(f"Cannot parse data to dataframe. Resource id: {resource.pk}; url: {url}")
        return None

    # Validate DGA file structure
    is_valid_file_structure: bool = validate_dga_df_columns(df)
    if not is_valid_file_structure:
        logger.error(f"Incorrect file structure. Resource id: {resource.pk}; url: {url}")
        return None

    return df


def create_main_dga_df(resources: QuerySet) -> pd.DataFrame:
    # List of columns from DGA Resource data shared with Main DGA DataFrame
    main_dga_columns: List[str] = [
        "Nazwa dysponenta zasobu",
        "Zasób chronionych danych",
        "Format danych",
        "Rozmiar danych",
    ]

    # Prepare empty Main DGA Resource DataFrame
    main_df: pd.DataFrame = pd.DataFrame(columns=main_dga_columns)

    # Create DataFrame for each Resource and concatenate it with Main DGA DataFrame
    count_dga_resources: int = resources.count()
    successful_resource_reads: int = 0
    for resource in resources:
        # Because CKAN resources' data are not stored in OD,
        # we have to create df based on currently available remote data
        if resource.is_imported_from_ckan:
            df: Optional[pd.DataFrame] = get_ckan_dga_resource_df(resource)
            if df is None:
                logger.error(f"Cannot read tabular data for CKAN harvested resource {resource.pk}")
                continue
            # Adjust DataFrame to Main DGA structure
            df["Nazwa dysponenta zasobu"] = np.nan  # will be filled later
            df = df[main_dga_columns]

        # Create Resource DataFrame for any other resource type
        else:
            # Read tabular data
            try:
                data = resource.tabular_data.table.read(keyed=True)
            except Exception as e:
                logger.error(f"Cannot read tabular data for for resource {resource.pk}: {e}")
                continue

            # Create and adjust DataFrame to Main DGA structure
            try:
                df: pd.DataFrame = pd.DataFrame(data, columns=main_df.columns)
            except Exception as e:
                logger.error(f"Cannot create DataFrame for resource {resource.pk}: {e}")
                sentry_sdk.api.capture_exception(e)
                continue

        # Clean and fill the data
        df = clean_columns_in_dataframe(df, "Zasób chronionych danych")
        institution: str = resource.institution.title
        df["Nazwa dysponenta zasobu"] = institution

        # Concatenate newly created DataFrame with Main DGA DataFrame
        main_df = pd.concat([main_df, df], ignore_index=True)
        successful_resource_reads += 1

    logger.info(f"Successful DGA Resources read: " f"{successful_resource_reads}/{count_dga_resources}.")

    if main_df.empty:
        logger.warning("Empty main DGA file.")

    # Insert a new column "Lp." at the first position in the DataFrame.
    # The column contains a sequence of numbers starting from 1 to the length
    # of the DataFrame.
    main_df.insert(0, "Lp.", range(1, len(main_df) + 1))

    # Add a new column "Warunki ponownego wykorzystywania" to the DataFrame.
    # Every row in this column is set to the string "określone w ofercie".
    main_df["Warunki ponownego wykorzystywania"] = "określone w ofercie"

    return main_df


def add_style_to_main_dga_excel_file(file_path: Path, sheet_name: str) -> None:
    """
    Stylizes the main DGA Excel file to enhance readability and presentation.

    This function applies several formatting adjustments to the Excel sheet:
    - Freezes the first row to keep headers visible while scrolling.
    - Removes all borders from cells.
    - Aligns all content to the left.
    - Sets fixed widths for the first six columns.

    Parameters:
    file_path (Path): The path to the Excel file that needs styling.
    sheet_name (str): The name of the sheet within the Excel file to be styled.

    The function modifies the Excel file in place, saving the changes directly
    to the provided file path.

    Note:
    - The function assumes that the sheet name provided exists in the Excel
      workbook.
    - Columns widths are set specifically for the first six columns as follows:
      Column 1: 6 units, Column 2: 32 units, Column 3: 95 units,
      Column 4: 16 units, Column 5: 16 units, Column 6: 36 units.

    See more: https://jira.coi.gov.pl/browse/OTD-689
    """
    workbook: Workbook = load_workbook(filename=file_path)
    worksheet: Worksheet = workbook[sheet_name]

    # Freeze first row
    worksheet.freeze_panes = "A2"

    # Remove all borders and align all content to left
    no_border: Border = Border(left=Side(style=None), right=Side(style=None), top=Side(style=None), bottom=Side(style=None))
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = no_border
            cell.alignment = Alignment(horizontal="left")

    # Set column widths
    columns_widths = {"A": 6, "B": 32, "C": 95, "D": 16, "E": 16, "F": 36}
    for col, width in columns_widths.items():
        worksheet.column_dimensions[col].width = width

    workbook.save(file_path)


@cache_memoize(
    timeout=settings.MAIN_DGA_RESOURCE_XLSX_CREATION_CACHE_TIMEOUT,
    hit_callable=lambda *args, **kwargs: logger.info("Using cache for main DGA file path."),
    key_generator_callable=key_generator_for_create_main_xlsx_file,
)
def create_main_dga_file() -> Path:
    """
    Creates the main DGA XLSX file based on all DGA resources.

    This function creates an XLSX file that contains information about all
    DGA resources. The file is saved to the file system, and the path to the
    created file is returned.

    Returns:
        Path: The path to the created XLSX file.
    """
    file_path: Path = get_or_create_main_dga_path()
    dga_resources: QuerySet = get_all_dga_resources_sorted_by_organizations()
    main_df: pd.DataFrame = create_main_dga_df(dga_resources)

    sheet_name: str = settings.MAIN_DGA_XLSX_WORKSHEET_NAME
    save_df_to_xlsx(df=main_df, file_path=file_path, sheet_name=sheet_name)
    add_style_to_main_dga_excel_file(file_path=file_path, sheet_name=sheet_name)
    return file_path


def check_all_resource_validations_status(resource: "Resource") -> None:  # noqa: F821
    """
    Checks the validation status of all tasks associated with the given
    resource.

    This function refreshes the resource from the database to ensure it has the
    latest data.
    It then checks the status of data tasks, file tasks, and link tasks
    associated with the resource.

    Args:
        resource (Resource): The resource object whose task statuses are to be
        checked.

    Raises:
        PendingValidationException: If any of the task validations are still
        pending.

        FailedValidationException: If any of the task validations are failed
        except for the validation of the empty file data task.
    """
    resource.refresh_from_db()
    data_task_status: str = resource.data_tasks_last_status
    file_task_status: str = resource.file_tasks_last_status
    link_task_status: str = resource.link_tasks_last_status

    statuses = {
        "data": data_task_status,
        "file": file_task_status,
        "link": link_task_status,
    }

    # All validations succeeded
    if all(status == states.SUCCESS for status in statuses.values()):
        logger.info(f"All validation for resource {resource.pk} succeeded.")
        return

    # Raise FailedValidation when file or link validation failed
    if any(status == states.FAILURE for status in [statuses["file"], statuses["link"]]):
        logger.error(
            f"Validation(s) for resource {resource.pk} failed: "
            f"data: {statuses['data']}; "
            f"file: {statuses['file']}; "
            f"link: {statuses['link']}."
        )
        raise FailedValidationException

    # Don't raise exception if the only reason for data validation failure
    # is an empty file
    if statuses["data"] == states.FAILURE:
        from mcod.resources.models import TaskResult

        data_task: Optional[TaskResult] = resource.data_tasks.last()
        data_failure_msg: Optional[List[str]] = data_task.message if data_task else None

        if data_failure_msg == [ZERO_DATA_ROWS_MSG]:
            if all(status == states.SUCCESS for status in [statuses["file"], statuses["link"]]):
                return
        else:
            logger.error(f"Data validation for resource {resource.pk} failed.")
            raise FailedValidationException

    # Raise PendingValidation exception in other scenarios
    logger.info(
        f"Pending validation(s) for resource {resource.pk}: "
        f"data: {statuses['data']}; "
        f"file: {statuses['file']}; "
        f"link: {statuses['link']}."
    )
    raise PendingValidationException


def get_default_main_dga_dataset_categories() -> List["Category"]:  # noqa: F821
    """
    Retrieves the main DGA dataset categories based on the titles specified in
    the settings. If a category with a specified title is not found, a warning
    is logged.
    """
    Category = apps.get_model("categories", "Category")
    categories_titles: List[str] = settings.MAIN_DGA_DATASET_CATEGORIES_TITLES

    found_categories: QuerySet = Category.objects.filter(title__in=categories_titles)
    found_titles: Set[str] = set(category.title for category in found_categories)

    missing_titles: Set[str] = set(categories_titles) - found_titles
    for missing_title in missing_titles:
        logger.warning(f"No category with title {missing_title} found.")

    return list(found_categories)


def get_or_create_default_main_dga_dataset_tags() -> List["Tag"]:  # noqa: F821
    """
    Retrieves or creates the main DGA dataset tags based on the names specified
    in the settings.
    """
    Tag = apps.get_model("tags", "Tag")
    tags_names: List[str] = settings.MAIN_DGA_DATASET_TAGS_NAMES
    pl_lang = settings.LANGUAGES[0][0]

    tags = []
    for tag_name in tags_names:
        tag, created = Tag.objects.get_or_create(name=tag_name, language=pl_lang)
        if created:
            logger.info(f"Created Tag with name {tag_name}.")
        tags.append(tag)

    return tags


def create_main_dga_dataset() -> int:
    """
    Creates the main DGA Dataset for the 'Ministerstwo Cyfryzacji' institution.

    Returns:
        int: The primary key of the newly created Dataset.

    Raises:
        ObjectDoesNotExist: If the Institution is not found.
    """
    Organization = apps.get_model("organizations", "Organization")
    Dataset = apps.get_model("datasets", "Dataset")
    Category = apps.get_model("categories", "Category")
    Tag = apps.get_model("tags", "Tag")

    # Get "Ministerstwo Cyfryzacji" institution object who is the owner of the
    # main DGA Dataset
    organization_pk: int = settings.MAIN_DGA_DATASET_OWNER_ORGANIZATION_PK
    try:
        institution: Organization = Organization.objects.get(pk=organization_pk)
    except Organization.DoesNotExist:
        logger.error(f"Can't create main DGA dataset. " f"Institution with pk {organization_pk} not found.")
        raise ObjectDoesNotExist("Main DGA Owner Institution not found.")

    dataset_params = {
        "title": settings.MAIN_DGA_DATASET_DEFAULT_TITLE,
        "notes": settings.MAIN_DGA_DATASET_DEFAULT_DESC,
        "organization": institution,
        "update_notification_recipient_email": settings.MAIN_DGA_DATASET_UPDATE_NOTIFICATION_EMAIL,
        "has_dynamic_data": False,
        "has_high_value_data": False,
        "has_high_value_data_from_ec_list": False,
        "has_research_data": False,
        "update_frequency": "daily",
        "status": "published",
    }

    with transaction.atomic():
        dataset = Dataset.objects.create(**dataset_params)

        categories: List[Category] = get_default_main_dga_dataset_categories()
        tags: List[Tag] = get_or_create_default_main_dga_dataset_tags()

        dataset.categories.set(categories)
        dataset.tags.set(tags)
        dataset.save()

    logger.info(f"Created main DGA dataset {dataset.pk}.")
    return dataset.pk


def key_generator_for_create_main_dga_resource(*args, **kwargs) -> str:
    return "main_dga_to_clean"


@cache_memoize(
    timeout=settings.MAIN_DGA_RESOURCE_CREATION_CACHE_TIMEOUT,
    hit_callable=lambda *args, **kwargs: logger.info("DGA objects get from cache."),
    key_generator_callable=key_generator_for_create_main_dga_resource,
)
def create_main_dga_resource_with_dataset(file_path: Path) -> Tuple[int, Optional[int]]:
    """
    Creates the main DGA Resource, including related ResourceFile objects and
    the main DGA Dataset for the `Ministerstwo Cyfryzacji` institution if it
    does not already exist.

    Args:
        file_path (Path): The path to the XLSX file to be used for the Resource.

    Returns:
        int: The primary key of the newly created Resource.
        Optional[int]: The Dataset primary key if created.
    """
    # Import models due to circular imports
    Resource = apps.get_model("resources", "Resource")
    ResourceFile = apps.get_model("resources", "ResourceFile")
    Dataset = apps.get_model("datasets", "Dataset")

    # Get info about current main DGA Resource and Dataset
    old_main_dga_resource: Optional[Resource] = get_main_dga_resource()
    main_dga_dataset: Optional[Dataset] = get_main_dga_dataset()

    # Create Resource and ResourceFile objects.
    new_main_dga_dataset_created: bool = False
    with transaction.atomic():
        # Create main DGA Dataset if it does not exist
        if main_dga_dataset is None:
            main_dga_dataset_pk: int = create_main_dga_dataset()
            new_main_dga_dataset_created = True
            main_dga_dataset: Dataset = Dataset.objects.get(pk=main_dga_dataset_pk)

        # Get metadata from existing resource or assign new one
        if old_main_dga_resource:
            title: str = old_main_dga_resource.title
            description: str = old_main_dga_resource.description
        else:
            title = settings.MAIN_DGA_RESOURCE_DEFAULT_TITLE
            description = settings.MAIN_DGA_RESOURCE_DEFAULT_DESC

        resource_params = {
            "title": title,
            "description": description,
            "dataset": main_dga_dataset,
            "has_dynamic_data": False,
            "has_high_value_data": False,
            "has_high_value_data_from_ec_list": False,
            "has_research_data": False,
            "contains_protected_data": True,
            "status": "published",
            "data_date": datetime.datetime.today(),
        }

        with open(file_path, "rb") as file:
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            file_name = os.path.basename(file.name)
            django_file = SimpleUploadedFile(name=file_name, content=file.read(), content_type=mimetype)
            resource = Resource.objects.create(**resource_params)
            logger.debug(f"Created Main DGA resource {resource.pk}")

            # this will run post_save signal which runs all validation tasks
            resource_file = ResourceFile.objects.create(resource=resource, is_main=True, file=django_file)
            logger.debug(f"Created Main DGA resource file {resource_file.pk}")

    return (
        resource.pk,
        main_dga_dataset.pk if new_main_dga_dataset_created else None,
    )


def update_or_create_aggr_dga_info_and_delete_old_main_dga(
    new_main_dga_resource: "Resource",  # noqa: F821
) -> None:
    """
    Updates or creates an entry in the AggregatedDGAInfo table with the new
    main DGA Resource.

    This function updates the existing entry in the AggregatedDGAInfo table
    to reference the new main DGA Resource. If no such entry exists,
    it creates a new one. Additionally, it deletes the old main DGA Resource
    to ensure that only one main DGA Resource is active at any given time.
    """
    Resource = apps.get_model("resources", "Resource")
    AggregatedDGAInfo = apps.get_model("resources", "AggregatedDGAInfo")
    dga_info: Optional[AggregatedDGAInfo] = AggregatedDGAInfo.objects.last()
    old_main_dga_resource: Optional[Resource] = dga_info.main_dga_resource if dga_info else None

    with transaction.atomic():
        if dga_info:
            dga_info.resource = new_main_dga_resource
            dga_info.views_count = new_main_dga_resource.dataset.computed_views_count
            dga_info.downloads_count = new_main_dga_resource.dataset.computed_downloads_count
            dga_info.save()
            logger.info(f"Updated AggregatedDGAInfo: {dga_info.pk} with resource: " f"{new_main_dga_resource.pk}.")

        else:
            dga_info: AggregatedDGAInfo = AggregatedDGAInfo.objects.create(resource=new_main_dga_resource)
            logger.info(f"Created AggregatedDGAInfo: {dga_info.pk} with resource: " f"{new_main_dga_resource.pk}.")

        if old_main_dga_resource:
            old_main_dga_resource.delete()
            logger.info(f"Previous Main DGA Resource " f"{old_main_dga_resource.pk} deleted.")


def clean_up_after_main_dga_resource_creation(exception_occurred: bool) -> None:
    """
    Cleans up objects created by the main_dga_resource_task only when it fails.
    This function will remove the Resource, ResourceFile and Dataset objects
    created during the task execution to maintain system integrity if
    exception_occurred is True. This function also remove created xlsx file
    from temp directory when task succeeded.
    """
    logger.info("Starting clean up process after main DGA Resource creation.")

    cache = caches["default"]
    clean_objects_key: str = key_generator_for_create_main_dga_resource()
    clean_file_path_key: str = key_generator_for_create_main_xlsx_file()

    resource_created: Optional[int]
    dataset_created: Optional[int]
    resource_created, dataset_created = cache.get(clean_objects_key, (None, None))
    file_path: Optional[str] = cache.get(clean_file_path_key, None)

    # delete created objects and release cache
    try:
        if exception_occurred and resource_created:
            Resource = apps.get_model("resources", "Resource")
            Resource.raw.filter(pk=resource_created).delete()

        if exception_occurred and dataset_created:
            Dataset = apps.get_model("datasets", "Dataset")
            Dataset.raw.filter(pk=dataset_created).delete()

        cache.delete(clean_objects_key)

    except Exception as exc:
        logger.error(f"Clean up failed. Reason: {exc}")
        sentry_sdk.api.capture_exception(exc)

    # delete file and release cache
    try:
        if file_path:
            os.remove(file_path)
            logger.info(f"File {file_path} has been deleted successfully.")
            cache.delete(clean_file_path_key)

    except Exception as e:
        logger.error(f"An error occurred while deleting the file {file_path}: {e}")

    logger.info("Clean up completed.")


def validate_contains_protected_data_with_other_metadata(
    contains_protected_data: bool,
    has_dynamic_data: Optional[bool],
    has_research_data: Optional[bool],
    has_high_value_data: Optional[bool],
    has_high_value_data_from_ec_list: Optional[bool],
) -> bool:

    if contains_protected_data and any(
        [has_dynamic_data, has_research_data, has_high_value_data, has_high_value_data_from_ec_list]
    ):
        return False
    return True


def validate_institution_type_for_contains_protected_data(contains_protected_data: bool, institution_type: str) -> bool:
    if contains_protected_data and institution_type not in ALLOWED_DGA_INSTITUTIONS:
        return False
    return True


def get_dga_resources_info_from_xml_harvester_file(loaded_data: List[Dict[str, Any]]) -> List[Tuple]:
    """
    Returns list of tuples with information about resources `extIdent` and `title`,
    which have `containsProtectedData` set as true.
    """
    dga_resources_info_from_xml: List[Tuple] = []
    for ds in loaded_data:
        for rs in ds.get("resources"):
            if rs.get("containsProtectedData"):
                resource_info = (rs.get("extIdent"), rs.get("title")["polish"])
                dga_resources_info_from_xml.append(resource_info)
    return dga_resources_info_from_xml


def request_remote_dga(url: str) -> requests.models.Response:
    """
    The function returns the response from the GET request to the remote DGA URL.
    It will raise requests.exceptions.RequestException when:
    - timeout will exceed,
    - other connection error will occur.
    """
    response = requests.get(url, verify=False, timeout=(3.0, 5.0))
    return response


def get_remote_extension_if_correct_dga_content_type(response: requests.models.Response) -> Optional[str]:
    """
    The function checks if the response corresponds to the content-type of the file with the extension appropriate for DGA
    and return extension.
    Otherwise, it will return None.
    """
    content_type: str = response.headers.get("Content-Type", "")
    extension_for_remote: Optional[str] = guess_extension(content_type)
    if extension_for_remote:
        extension_for_remote = extension_for_remote[1:]
        if extension_for_remote in DGA_RESOURCE_EXTENSIONS:
            return extension_for_remote
    return None
